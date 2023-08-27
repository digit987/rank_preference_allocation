from openpyxl import load_workbook

file_path=""

workbook = load_workbook(file_path)

sheet = workbook['REVISED MAIN']

total_rows=sheet.max_row
total_cols=sheet.max_column

excel_to_list = []

'''
Depts are represented by following indices
0: ndmc,
1: edmc,
2: damb,
3: dusib,
4: dtl,
5: dsiidc,
6: ifc,
7: dtc
'''

#Mapping department name to index
dept_to_index = {
        "NDMC": 0,
        "EDMC": 1,
        "DAMB": 2,
        "DUSIB": 3,
        "DTL": 4,
        "DSIIDC": 5,
        "I&FC": 6,
        "DTC": 7
    }

#Mapping index to department name
index_to_dept = {
        0: "NDMC",
        1: "EDMC",
        2: "DAMB",
        3: "DUSIB",
        4: "DTL",
        5: "DSIIDC",
        6: "I&FC",
        7: "DTC" 
    }

dept_and_vacancies = [

{
        "ur": 50,
        "obc(d)": 25,
        "ews": 11,
        "sc": 12,
        "st": 12,
        "ph_vh": 4,
        "ph_oh":3,
        "ph_hh": 3,
        "exsm": 3,
        "sportsperson": 2
    },

{
        "ur": 50,
        "obc(d)": 25,
        "ews": 11,
        "sc": 12,
        "st": 12,
        "ph_vh": 4,
        "ph_oh":3,
        "ph_hh": 3,
        "exsm": 3,
        "sportsperson": 2
    },

{
        "ur": 50,
        "obc(d)": 25,
        "ews": 11,
        "sc": 12,
        "st": 12,
        "ph_vh": 4,
        "ph_oh":3,
        "ph_hh": 3,
        "exsm": 3,
        "sportsperson": 2
    },

{
        "ur": 50,
        "obc(d)": 25,
        "ews": 11,
        "sc": 12,
        "st": 12,
        "ph_vh": 4,
        "ph_oh":3,
        "ph_hh": 3,
        "exsm": 3,
        "sportsperson": 2
    },

{
        "ur": 50,
        "obc(d)": 25,
        "ews": 11,
        "sc": 12,
        "st": 12,
        "ph_vh": 4,
        "ph_oh":3,
        "ph_hh": 3,
        "exsm": 3,
        "sportsperson": 2
    },

{
        "ur": 50,
        "obc(d)": 25,
        "ews": 11,
        "sc": 12,
        "st": 12,
        "ph_vh": 4,
        "ph_oh":3,
        "ph_hh": 3,
        "exsm": 3,
        "sportsperson": 2
    },

{
        "ur": 50,
        "obc(d)": 25,
        "ews": 11,
        "sc": 12,
        "st": 12,
        "ph_vh": 4,
        "ph_oh":3,
        "ph_hh": 3,
        "exsm": 3,
        "sportsperson": 2
    },
{
        "ur": 50,
        "obc(d)": 25,
        "ews": 11,
        "sc": 12,
        "st": 12,
        "ph_vh": 4,
        "ph_oh":3,
        "ph_hh": 3,
        "exsm": 3,
        "sportsperson": 2
    }
]

#Reading Excel file row by row
for i in range(945):
    each_row = []
    each_row.append(sheet.cell(row = i+1, column = 1).value) #Serial number
    each_row.append(sheet.cell(row = i+1, column = 2).value) #Application Number
    each_row.append(sheet.cell(row = i+1, column = 7).value) #Preference 1
    each_row.append(sheet.cell(row = i+1, column = 8).value) #Preference 2
    each_row.append(sheet.cell(row = i+1, column = 9).value) #Preference 3
    each_row.append(sheet.cell(row = i+1, column = 10).value) #Preference 4
    each_row.append(sheet.cell(row = i+1, column = 11).value) #Preference 5
    each_row.append(sheet.cell(row = i+1, column = 12).value) #Preference 6
    each_row.append(sheet.cell(row = i+1, column = 13).value) #Preference 7
    each_row.append(sheet.cell(row = i+1, column = 14).value) #Preference 8
    each_row.append(sheet.cell(row = i+1, column = 15).value) #Category
    each_row.append(sheet.cell(row = i+1, column = 16).value) #Sub category
    each_row.append(sheet.cell(row = i+1, column = 19).value) #Marks
    excel_to_list.append(each_row)
    
#print(excel_to_list)    

print("==================================PH_VH allocation============================================")
for row in range(len(excel_to_list)):
    candidate_details = excel_to_list[row]
    category = candidate_details[10]
    if category == "UR":
        category = "ur"
    if category == "OBC(D)":
        category = "obc(d)"
    if category == "EWS":
        category = "ews"
    if category == "SC":
        category = "sc"
    if category == "ST":
        category = "st"
    subcategory = candidate_details[11]
    if candidate_details[11] == "PH-VH":
        preference_list = candidate_details[2:10]
        preference_iterator = 0
        while preference_iterator < len(preference_list): 
            if preference_list[preference_iterator] is not None:
                print("Candidate", candidate_details[0], "has category", category, "and subcategory PH_VH and has given Preference", preference_iterator+1, "to department", preference_list[preference_iterator][3:])
                #ph_vh remaining vacancies in a particular preferred department
                if dept_and_vacancies[dept_to_index[preference_list[preference_iterator][3:]]]["ph_vh"] > 0:
                    dept_and_vacancies[dept_to_index[preference_list[preference_iterator][3:]]]["ph_vh"] -= 1
                    dept_and_vacancies[dept_to_index[preference_list[preference_iterator][3:]]][category] -= 1
                    break
            preference_iterator += 1
        print()

for i in range(len(dept_and_vacancies)):
    print("Remaining vacancies in department", index_to_dept[i], "are:", dept_and_vacancies[i])
print("\n\n\n\n\n\n")
    

print("====================================PH_OH allocation===========================================")
for row in range(len(excel_to_list)):
    candidate_details = excel_to_list[row]
    category = candidate_details[10]
    if category == "UR":
        category = "ur"
    if category == "OBC(D)":
        category = "obc(d)"
    if category == "EWS":
        category = "ews"
    if category == "SC":
        category = "sc"
    if category == "ST":
        category = "st"
    subcategory = candidate_details[11]
    if candidate_details[11] == "PH-OH":
        preference_list = candidate_details[2:10]
        preference_iterator = 0
        while preference_iterator < len(preference_list): 
            if preference_list[preference_iterator] is not None:
                print("Candidate", candidate_details[0], "has category", category, "and subcategory PH_OH and has given Preference", preference_iterator+1, "to department", preference_list[preference_iterator][3:])
                #ph_oh remaining vacancies in a particular preferred department
                if dept_and_vacancies[dept_to_index[preference_list[preference_iterator][3:]]]["ph_oh"] > 0:
                    dept_and_vacancies[dept_to_index[preference_list[preference_iterator][3:]]]["ph_oh"] -= 1
                    dept_and_vacancies[dept_to_index[preference_list[preference_iterator][3:]]][category] -= 1
                    break
            preference_iterator += 1
        print()

for i in range(len(dept_and_vacancies)):
    print("Remaining vacancies in department", index_to_dept[i], "are:", dept_and_vacancies[i])
print("\n\n\n\n\n\n")

print("==================================PH_HH allocation=============================================")
for row in range(len(excel_to_list)):
    candidate_details = excel_to_list[row]
    category = candidate_details[10]
    if category == "UR":
        category = "ur"
    if category == "OBC(D)":
        category = "obc(d)"
    if category == "EWS":
        category = "ews"
    if category == "SC":
        category = "sc"
    if category == "ST":
        category = "st"
    subcategory = candidate_details[11]
    if candidate_details[11] == "PH-HH":
        preference_list = candidate_details[2:10]
        preference_iterator = 0
        while preference_iterator < len(preference_list): 
            if preference_list[preference_iterator] is not None:
                print("Candidate", candidate_details[0], "has category", category, "and subcategory PH_HH and has given Preference", preference_iterator+1, "to department", preference_list[preference_iterator][3:])
                #ph_oh remaining vacancies in a particular preferred department
                if dept_and_vacancies[dept_to_index[preference_list[preference_iterator][3:]]]["ph_hh"] > 0:
                    dept_and_vacancies[dept_to_index[preference_list[preference_iterator][3:]]]["ph_hh"] -= 1
                    dept_and_vacancies[dept_to_index[preference_list[preference_iterator][3:]]][category] -= 1
                    break
            preference_iterator += 1
        print()

for i in range(len(dept_and_vacancies)):
    print("Remaining vacancies in department", index_to_dept[i], "are:", dept_and_vacancies[i])
print("\n\n\n\n\n\n")

print("==================================EXSM allocation================================================")
for row in range(len(excel_to_list)):
    candidate_details = excel_to_list[row]
    category = candidate_details[10]
    if category == "UR":
        category = "ur"
    if category == "OBC(D)":
        category = "obc(d)"
    if category == "EWS":
        category = "ews"
    if category == "SC":
        category = "sc"
    if category == "ST":
        category = "st"
    subcategory = candidate_details[11]
    if candidate_details[11] == "EXSM":
        preference_list = candidate_details[2:10]
        preference_iterator = 0
        while preference_iterator < len(preference_list): 
            if preference_list[preference_iterator] is not None:
                print("Candidate", candidate_details[0], "has category", category, "and subcategory EXSM and has given Preference", preference_iterator+1, "to department", preference_list[preference_iterator][3:])
                #ph_oh remaining vacancies in a particular preferred department
                if dept_and_vacancies[dept_to_index[preference_list[preference_iterator][3:]]]["exsm"] > 0:
                    dept_and_vacancies[dept_to_index[preference_list[preference_iterator][3:]]]["exsm"] -= 1
                    dept_and_vacancies[dept_to_index[preference_list[preference_iterator][3:]]][category] -= 1
                    break
            preference_iterator += 1
        print()

for i in range(len(dept_and_vacancies)):
    print("Remaining vacancies in department", index_to_dept[i], "are:", dept_and_vacancies[i])
print("\n\n\n\n\n\n")

print("===================================Sportsperson Allocation=========================================")
for row in range(len(excel_to_list)):
    candidate_details = excel_to_list[row]
    category = candidate_details[10]
    if category == "UR":
        category = "ur"
    if category == "OBC(D)":
        category = "obc(d)"
    if category == "EWS":
        category = "ews"
    if category == "SC":
        category = "sc"
    if category == "ST":
        category = "st"
    subcategory = candidate_details[11]
    if candidate_details[11] == "sportsperson":
        preference_list = candidate_details[2:10]
        preference_iterator = 0
        while preference_iterator < len(preference_list): 
            if preference_list[preference_iterator] is not None:
                print("Candidate", candidate_details[0], "has category", category, "and subcategory Sportsperson and has given Preference", preference_iterator+1, "to department", preference_list[preference_iterator][3:])
                #ph_oh remaining vacancies in a particular preferred department
                if dept_and_vacancies[dept_to_index[preference_list[preference_iterator][3:]]]["sportsperson"] > 0:
                    dept_and_vacancies[dept_to_index[preference_list[preference_iterator][3:]]]["sportsperson"] -= 1
                    dept_and_vacancies[dept_to_index[preference_list[preference_iterator][3:]]][category] -= 1
                    break
            preference_iterator += 1
        print()

for i in range(len(dept_and_vacancies)):
    print("Remaining vacancies in department", index_to_dept[i], "are:", dept_and_vacancies[i], "\n")
